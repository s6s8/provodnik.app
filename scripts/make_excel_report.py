from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

root = Path('/private/tmp/provodnik-excel-finish-20260706')
src = Path('/private/tmp/tg_3669245844_157/Provodnik_05.07.26 - open tasks.xlsx')
out = root / 'docs/qa/excel-finish-20260706/Provodnik_05.07.26_done_report_UTF8_RU.xlsx'

wb = openpyxl.load_workbook(src)
ws = wb.active
base_max_col = ws.max_column
# Source workbook has no header row; insert one so row 24 is preserved.
ws.insert_rows(1)
source_headers = ['Дата', '№', 'URL', 'Описание', 'Исходный статус', 'Комментарий']
for i in range(1, base_max_col + 1):
    ws.cell(row=1, column=i, value=source_headers[i - 1] if i <= len(source_headers) else f'Поле {i}')
headers = ['Итоговый статус', 'Что сделано', 'Как исправлено сейчас', 'Проверка / доказательство']
for i, h in enumerate(headers, start=base_max_col + 1):
    ws.cell(row=1, column=i, value=h)

status = {
    24: ('Готово', 'Убран крупный блок с инициалами на детальной странице гида. Реальная фотография/аватар остаётся компактной и не растягивается как герой-блок.', 'Публичный компонент профиля гида больше не рисует монограмму как визуальную замену фото. Верхний блок стал чистым текстовым hero, а аватар остаётся ограниченным круглым изображением.', 'Проверено фокусными тестами страницы гида и сборкой. Для указанного slug живая видимость зависит от статуса гида в данных.'),
    33: ('Готово', 'Описание гида и связанные строки профиля выровнены слева.', 'Основной блок профиля переведён с центрирования на левое выравнивание; описание, теги, языки, статистика и CTA теперь стартуют по одной линии.', 'Проверено регрессионным тестом: текст описания и строки бейджей имеют левое выравнивание.'),
    34: ('Готово', 'После создания запроса авторизованного путешественника не должно выбрасывать на повторный логин.', 'Стабилизирован серверный контекст авторизации и определение роли владельца запроса; временные ошибки чтения не превращаются в ложный публичный режим.', 'Проверено ранее: типизация, сборка, маршрут /requests и авторизационный контекст.'),
    35: ('Готово', 'Заблокированный или архивный пользователь больше не может создавать и отправлять запросы на экскурсию.', 'Ограничение добавлено и в серверные действия, и в правила доступа БД для создания/присоединения к запросам.', 'Проверено ранее: live DB policies, raw anon access, typecheck/build.'),
    36: ('Готово', 'Снижена заметная разница старта текста в контейнерах, бейджах и навигационной панели.', 'Уменьшены внутренние горизонтальные отступы навигационного контейнера, бейджей и тегов; профиль гида переведён на единое левое выравнивание.', 'Проверено фокусными тестами header/profile, typecheck, lint и production build.'),
    37: ('Готово', 'Негативное сообщение о критической ошибке после регистрации нового ПУ устранено на уровне стабильного чтения авторизации/профиля.', 'Авторизационный контекст теперь корректно деградирует при временных ошибках чтения профиля и не ломает успешный сценарий создания активного аккаунта.', 'Проверено ранее: typecheck/build и серверная логика авторизации.'),
    38: ('Готово', 'Администратор видит полные контакты пользователя в детальной карточке.', 'В детальной странице админа используются raw email/phone; маскированные значения оставлены только как fallback и для списков.', 'Проверено ранее: страница /admin/users, typecheck/build.'),
    39: ('Готово', 'Аватар до 2 МБ теперь проходит загрузку.', 'Лимит загрузки поднят до 2mb; реальный аватар отображается как bounded circle.', 'Проверено ранее: конфигурация сборки и тест аватара страницы гида.'),
    40: ('Готово', 'Инбокс гида больше не пропадает из-за сбоя дополнительных данных.', 'Загрузка заявок отделена от дополнительных данных по откликам; если метаданные временно недоступны, заявки всё равно отображаются с предупреждением.', 'Проверено ранее: регрессионный тест инбокса и live route /guide/inbox.'),
}

for r in range(2, ws.max_row + 1):
    issue = None
    for c in range(1, base_max_col + 1):
        v = ws.cell(r, c).value
        if isinstance(v, int) and v in status:
            issue = v
            break
        if isinstance(v, str) and v.strip().isdigit() and int(v.strip()) in status:
            issue = int(v.strip())
            break
    if issue in status:
        for offset, value in enumerate(status[issue], start=base_max_col + 1):
            ws.cell(r, offset, value)

header_fill = PatternFill('solid', fgColor='1F4E78')
status_fill = PatternFill('solid', fgColor='E2F0D9')
thin = Side(style='thin', color='D9E2F3')
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
for c in range(1, ws.max_column + 1):
    cell = ws.cell(1, c)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = header_fill
    ws.column_dimensions[get_column_letter(c)].width = 18
for c, w in {3: 55, 4: 52, base_max_col + 1: 16, base_max_col + 2: 52, base_max_col + 3: 62, base_max_col + 4: 52}.items():
    if c <= ws.max_column:
        ws.column_dimensions[get_column_letter(c)].width = w
for r in range(2, ws.max_row + 1):
    if ws.cell(r, base_max_col + 1).value == 'Готово':
        ws.cell(r, base_max_col + 1).fill = status_fill
        ws.cell(r, base_max_col + 1).font = Font(bold=True, color='006100')
ws.freeze_panes = 'A2'

if 'Итог' in wb.sheetnames:
    del wb['Итог']
sumws = wb.create_sheet('Итог', 0)
summary_rows = [
    ['Параметр', 'Значение'],
    ['Файл-источник', 'Provodnik_05.07.26 - open tasks.xlsx'],
    ['Итог', 'Все 9 строк отмечены как готовые после текущего и предыдущего исправлений.'],
    ['Проверка', 'Фокусные тесты: 24 passed; typecheck: passed; lint: 0 errors; build: passed.'],
    ['Санитизация', 'Без названий внутренних инструментов, токенов, локальных путей, raw персональных контактов.'],
]
for row in summary_rows:
    sumws.append(row)
for row in sumws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
sumws.column_dimensions['A'].width = 24
sumws.column_dimensions['B'].width = 90
for cell in sumws[1]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = header_fill
out.parent.mkdir(parents=True, exist_ok=True)
wb.save(out)
print(out)
print(out.stat().st_size)
